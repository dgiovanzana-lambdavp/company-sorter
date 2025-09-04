from flask import Flask, request, render_template_string, jsonify, make_response
import os, io, csv, json
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)

# -------------------- State persistence --------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_FILE = os.path.join(BASE_DIR, "state.json")
all_rows: list[list[str]] = []
headers: list[str] = []
keep_rows: list[list[str]] = []
skip_rows: list[list[str]] = []
current_record: dict[str, str] | None = None

# -------------------- Persistence Helpers --------------------

def save_state() -> None:
    """Write the current state to disk."""
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "all_rows": all_rows,
            "headers": headers,
            "keep_rows": keep_rows,
            "skip_rows": skip_rows,
            "current_record": current_record,
        }, f)


def load_state() -> None:
    """Load state from disk if available."""
    global all_rows, headers, keep_rows, skip_rows, current_record
    if not os.path.exists(STATE_FILE):
        return
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            s = json.load(f)
            all_rows[:] = s.get("all_rows", [])
            headers[:] = s.get("headers", [])
            keep_rows[:] = s.get("keep_rows", [])
            skip_rows[:] = s.get("skip_rows", [])
            current_record = s.get("current_record")
    except Exception:
        pass

# right after load_state() definition
@app.before_request
def restore_state():
    load_state()

# -------------------- HTML Template --------------------
TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Contacts Editor & Sorter</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 2rem; }
    .hidden { display: none !important; }
    #dragArea { border: 2px dashed #888; padding: 2rem; text-align: center; color: #555; cursor: pointer; }
    #dragArea.dragover { background: #eef; }
    #status { margin-top: 1rem; font-weight: bold; }
    /* make all three buttons match */
    #openModalBtn,
    #downloadBtn,
    #deleteStateBtn {
      margin: 0.5rem 0;
      padding: 0.5rem 1rem;
      border: none;
      border-radius: 4px;
      cursor: pointer;
    }
    #openModalBtn {
      background: #4CAF50;
      color: white;
    }
    #downloadBtn {
      background: #2196F3;
      color: white;
    }
    #deleteStateBtn {
      background: #d9534f;
      color: white;
    }

    #fileInput {
      position: absolute;
      left: -9999px;
      top: -9999px;
      opacity: 0;
    }
    #lists { margin-top: 1rem; display: flex; gap: 2rem; }
    #lists > div { flex: 1; }
    #lists ul { list-style: none; padding: 0.5rem; border: 1px solid #ccc; max-height: 150px; overflow: auto; }
    .modal-backdrop { position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.5); display: none; align-items: center; justify-content: center; z-index: 1000; }
    .modal { background: white; padding: 1rem; border-radius: 8px; max-width: 600px; width: 90%; max-height: 80%; display: flex; flex-direction: column; position: relative; }
    .modal-header { font-size: 1.2rem; margin-bottom: 1rem; }
    .modal-body { flex: 1; overflow: auto; }
    .modal-footer { display: flex; justify-content: space-between; margin-top: 1rem; }
    #skipBtn { background: #F4D03F; color: black; border: none; border-radius: 4px; padding: 0.5rem 1rem; cursor: pointer; }
    #keepBtn { background: #E44E3C; color: white; border: none; border-radius: 4px; padding: 0.5rem 1rem; cursor: pointer; }
    .field { display: flex; margin-bottom: 0.8rem; }
    .field label { width: 150px; font-weight: bold; }
    .field input { flex: 1; padding: 0.3rem; }
    #loading { display: none; color: #6A1B9A; }
    /* Make list items highlightable and clickable */
    #lists li.clickable:hover {
      background-color: #eef;
      cursor: pointer;
    }  /* ← close the rule */

    /* side buttons */
    #searchCeoBtn, #searchWebsiteBtn, #openWebsiteBtn, #openLinkedinBtn {
      position: absolute;
      left: 100%;
      margin-left: 10px;
      background: #007bff;
      color: white;
      border: none;
      padding: 0.5rem 1rem;
      border-radius: 4px;
      cursor: pointer;
      opacity: 0.8;
      white-space: nowrap;
    }
    #searchCeoBtn { top: 0; }
    #searchWebsiteBtn { top: 40px; }
    #openWebsiteBtn { top: 80px; }
    #openLinkedinBtn { top: 120px; }



  </style>
</head>
<body>
  <h2>Contacts Editor & Sorter</h2>
  <div id="dragArea">Drag or click to upload .csv or .xlsx</div>
  <input id="fileInput" type="file" accept=".csv,.xlsx">
  <p id="status"></p>
  <button id="openModalBtn" class="hidden">Open Sorting Window</button>
  <button id="downloadBtn" class="hidden">Download Kept Records</button>
  <button id="deleteStateBtn" class="hidden" style="background-color:#d9534f;color:white;">Delete State</button>
  <div id="lists">
    <div><h3>Kept Records</h3><ul id="keptList"></ul></div>
    <div><h3>Skipped Records</h3><ul id="skippedList"></ul></div>
  </div>
  <div id="loading">Loading…</div>

  <div id="modalBackdrop" class="modal-backdrop">
    <div class="modal">
      <button id="searchCeoBtn">Search CEO LinkedIn</button>
      <button id="searchWebsiteBtn">Search Website</button>
      <button id="openWebsiteBtn">Open Website [Enter]</button>
      <button id="openLinkedinBtn">Open LinkedIn [Q]</button>
      <div class="modal-header">Edit & Decide</div>
      <div class="modal-body" id="recordFields"></div>
      <div class="modal-footer">
        <button id="skipBtn">Skip [E]</button>
        <button id="keepBtn">Keep [J]</button>
      </div>
    </div>
  </div>

  <script>
  // —— DOM refs —— 
  const dragArea       = document.getElementById('dragArea');
  const fileInput      = document.getElementById('fileInput');
  const statusP        = document.getElementById('status');
  const openBtn        = document.getElementById('openModalBtn');
  const downloadBtn    = document.getElementById('downloadBtn');
  const deleteStateBtn = document.getElementById('deleteStateBtn');
  const keptList       = document.getElementById('keptList');
  const skippedList    = document.getElementById('skippedList');
  const modal          = document.getElementById('modalBackdrop');
  const recordFields   = document.getElementById('recordFields');
  const skipBtn        = document.getElementById('skipBtn');
  const keepBtn        = document.getElementById('keepBtn');
  const loading        = document.getElementById('loading');
  const searchCeoBtn   = document.getElementById('searchCeoBtn');
  const searchWebsiteBtn   = document.getElementById('searchWebsiteBtn');
  const openWebsiteBtn     = document.getElementById('openWebsiteBtn');
  const openLinkedinBtn  = document.getElementById('openLinkedinBtn');

  function openLinkedin() {                                              // new
  if (current && current['Linkedin']) {
    let u = current['Linkedin'];
    if (!/^https?:\/\//i.test(u)) u = 'http://' + u;
    window.open(u, '_blank');
  }
  }
  openLinkedinBtn.addEventListener('click', openLinkedin);  

  let headers = [], current = null, originalRow = null;

  // —— Restore persisted state on load —— 
  window.addEventListener('load', function() {
    fetch('/get_state')
      .then(r => r.json())
      .then(d => {
        if (d.total || d.kept || d.skipped) {
          headers = d.headers;
          renderLists(d.keep_rows, d.skip_rows);
          statusP.textContent = `State: ${d.total} left, ${d.kept} kept, ${d.skipped} skipped.`;
          openBtn.classList.remove('hidden');
          downloadBtn.classList.remove('hidden');
          deleteStateBtn.classList.remove('hidden');
        }
      })
      .catch(console.error);
  });

  // —— Upload handler —— 
  function handleUpload() {
    if (!fileInput.files.length) return;
    statusP.textContent = 'Uploading…';
    const fd = new FormData();
    fd.append('file', fileInput.files[0]);
    fetch('/upload', { method: 'POST', body: fd })
      .then(r => r.json())
      .then(d => {
        if (d.error) {
          statusP.textContent = d.error;
          return;
        }
        headers = d.headers;
        statusP.textContent = `Loaded ${d.totalRecords} records.`;
        renderLists(d.keep_rows, d.skip_rows);
        openBtn.classList.remove('hidden');
        downloadBtn.classList.remove('hidden');
        deleteStateBtn.classList.remove('hidden');
      })
      .catch(() => statusP.textContent = 'Upload failed');
  }

  // —— Drag & drop / click to upload —— 
  ['dragenter', 'dragover'].forEach(ev =>
    dragArea.addEventListener(ev, e => {
      e.preventDefault();
      dragArea.classList.add('dragover');
    })
  );
  ['dragleave', 'drop'].forEach(ev =>
    dragArea.addEventListener(ev, e => {
      e.preventDefault();
      dragArea.classList.remove('dragover');
    })
  );
  dragArea.addEventListener('drop', e => {
    fileInput.files = e.dataTransfer.files;
    handleUpload();
  });
  dragArea.addEventListener('click', () => fileInput.click());
  fileInput.addEventListener('change', handleUpload);

  // —— Delete State button —— 
  deleteStateBtn.addEventListener('click', () => {
    if (!confirm('Really delete state?')) return;
    fetch('/delete_state', { method: 'POST' })
      .then(r => r.json())
      .then(d => {
        statusP.textContent = d.message;
        openBtn.classList.add('hidden');
        downloadBtn.classList.add('hidden');
        deleteStateBtn.classList.add('hidden');
        keptList.innerHTML = skippedList.innerHTML = '';
      });
  });

  // —— Render kept/skipped lists —— 
  function renderLists(kept, skipped) {
    keptList.innerHTML = skippedList.innerHTML = '';
    const ci = headers.indexOf('Company');
    function makeItem(r) {
      const li = document.createElement('li');
      li.textContent = (ci >= 0 ? r[ci] : r[0]) || '—';
      li.classList.add('clickable');
      li.addEventListener('click', () => {
        current = {};
        headers.forEach((h, i) => current[h] = r[i] || '');
        originalRow = r.slice();
        showFields();
        modal.style.display = 'flex';
      });
      return li;
    }
    kept.forEach(r => keptList.appendChild(makeItem(r)));
    skipped.forEach(r => skippedList.appendChild(makeItem(r)));
  }

  // —— Modal logic —— 
  openBtn.addEventListener('click', () => {
    modal.style.display = 'flex';
    nextRecord();
  });
  modal.addEventListener('click', e => {
    if (e.target === modal) modal.style.display = 'none';
  });

  function nextRecord() {
    loading.style.display = 'block';
    fetch('/next_record')
      .then(r => r.json())
      .then(d => {
        loading.style.display = 'none';
        if (d.done) {
          recordFields.innerHTML = '<p>No more records.</p>';
          current = null;
          originalRow = null;
        } else {
          current = d.row;
          originalRow = null;
          showFields();
        }
      });
  }

  function showFields() {
    recordFields.innerHTML = '';
    headers.forEach(h => {
      const div = document.createElement('div');
      div.className = 'field';
      const lbl = document.createElement('label');
      lbl.textContent = h;
      const inp = document.createElement('input');
      inp.value = current[h] || '';
      inp.dataset.field = h;
      div.append(lbl, inp);
      recordFields.appendChild(div);
    });
    // no auto-focus, so Space can be intercepted below
  }

  function collectRow() {
    const obj = {};
    recordFields
      .querySelectorAll('input[data-field]')
      .forEach(i => obj[i.dataset.field] = i.value);
    return obj;
  }

  function decide(url) {
    fetch(url, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ row: collectRow(), original: originalRow })
    })
    .then(r => r.json())
    .then(d => {
      renderLists(d.keep_rows, d.skip_rows);
      nextRecord();
    });
  }

  skipBtn.addEventListener('click', () => decide('/skip_record'));
  keepBtn.addEventListener('click', () => decide('/keep_record'));
  downloadBtn.addEventListener('click', () => window.location = '/download');

  searchCeoBtn.addEventListener('click', () => {
    const companyInput = recordFields.querySelector('input[data-field="Company"]');
    if (companyInput && companyInput.value.trim()) {
      const query = encodeURIComponent(companyInput.value.trim() + ' CEO LinkedIn');
      window.open(`https://www.google.com/search?q=${query}`, '_blank');
    }
  });

  searchWebsiteBtn.addEventListener('click', () => {
    const fields = ['Company', 'Address', 'City', 'State', 'Zip'];
    const parts = fields.map(field => {
      const inp = recordFields.querySelector(`input[data-field="${field}"]`);
      return inp ? inp.value.trim() : '';
    }).filter(part => part !== '');
    if (parts.length > 0) {
      const query = encodeURIComponent(parts.join(' '));
      window.open(`https://www.google.com/search?q=${query}`, '_blank');
    }
  });

  openWebsiteBtn.addEventListener('click', () => {
  if (current && current['Website']) {
    let u = current['Website'];
    if (!/^https?:\/\//i.test(u)) u = 'http://' + u;
    window.open(u, '_blank');
  }
});

  // —— Keyboard handling —— 
  document.addEventListener('keydown', function(e) {
    // 1) closed + Enter → open modal
    if (modal.style.display !== 'flex' && e.key === 'Enter') {
      openBtn.click();
      e.preventDefault();
      return;
    }
    // 2) open + no focus + Space → focus "Compliment"
    if (
      modal.style.display === 'flex' &&
      document.activeElement.tagName !== 'INPUT' &&
      e.key === ' '
    ) {
      const c = recordFields.querySelector('input[data-field="Compliment"]');
      if (c) c.focus();
      e.preventDefault();
      return;
    }
    // 3) open + no focus + Enter → open website
    if (
      modal.style.display === 'flex' &&
      document.activeElement.tagName !== 'INPUT' &&
      e.key === 'Enter'
    ) {
      if (current && current['Website']) {
        let u = current['Website'];
        if (!/^https?:\/\//i.test(u)) u = 'http://' + u;
        window.open(u, '_blank');
      }
      e.preventDefault();
      return;
    }
    // 3b) open + no focus + Q → open LinkedIn
    if (
      modal.style.display === 'flex' &&
      document.activeElement.tagName !== 'INPUT' &&
      (e.code === 'KeyQ' || e.key === 'q' || e.key === 'Q')
    ) {
      openLinkedin();           // calls the function you added in step 3
      e.preventDefault();
      return;
    }
    // 4) Escape → blur or close
    if (e.key === 'Escape') {
      if (document.activeElement.tagName === 'INPUT') {
        document.activeElement.blur();
      } else if (modal.style.display === 'flex') {
        modal.style.display = 'none';
      }
      e.preventDefault();
      return;
    }
    // 5) open + no focus + E/J → skip/keep
    if (
      modal.style.display === 'flex' &&
      document.activeElement.tagName !== 'INPUT' &&
      e.key.toLowerCase() === 'e'
    ) {
      skipBtn.click();
      e.preventDefault();
    }
    if (
      modal.style.display === 'flex' &&
      document.activeElement.tagName !== 'INPUT' &&
      e.key.toLowerCase() === 'j'
    ) {
      keepBtn.click();
      e.preventDefault();
    }
  });
</script>

</body>
</html>
"""

# -------------------- Flask Endpoints --------------------

@app.route('/')
def index():
    return render_template_string(TEMPLATE)

@app.route('/upload', methods=['POST'])
def upload():
    f = request.files.get('file')
    if not f:
        return jsonify(error='No file uploaded'), 400
    name = f.filename.lower()
    global headers, all_rows, keep_rows, skip_rows, current_record
    headers.clear(); all_rows.clear(); keep_rows.clear(); skip_rows.clear(); current_record = None
    data = f.read()
    if name.endswith('.csv'):
        text = data.decode('utf-8', errors='ignore').splitlines()
        reader = csv.reader(text)
        for i, row in enumerate(reader):
            if i == 0:
                headers.extend(row)
            else:
                all_rows.append(row)
    elif name.endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        for i, rw in enumerate(wb.active.iter_rows(values_only=True)):
            if i == 0:
                headers.extend([str(c) for c in rw if c is not None])
            else:
                all_rows.append([str(c) if c is not None else '' for c in rw])
    else:
        return jsonify(error='Unsupported file type'), 400
    save_state()
    return jsonify(headers=headers, totalRecords=len(all_rows), keep_rows=keep_rows, skip_rows=skip_rows)

@app.route('/get_state')
def get_state():
    return jsonify(total=len(all_rows), kept=len(keep_rows), skipped=len(skip_rows), headers=headers, keep_rows=keep_rows, skip_rows=skip_rows)

@app.route('/next_record')
def next_record():
    global current_record
    if current_record is not None:
        return jsonify(done=False, row=current_record)
    if not all_rows:
        return jsonify(done=True)
    row = all_rows.pop(0)
    current_record = dict(zip(headers, row))
    save_state()
    return jsonify(done=False, row=current_record)

@app.route('/skip_record', methods=['POST'])
def skip_record():
    global current_record
    data = request.json
    row = data.get('row', {})
    orig = data.get('original')
    row_list = [row.get(h, '') for h in headers]
    if orig is not None:
        if orig in keep_rows:
            keep_rows.remove(orig)
        if orig in skip_rows:
            skip_rows.remove(orig)
    if row_list in keep_rows:
        keep_rows.remove(row_list)
    skip_rows.append(row_list)
    current_record = None
    save_state()
    return jsonify(keep_rows=keep_rows, skip_rows=skip_rows)

@app.route('/keep_record', methods=['POST'])
def keep_record():
    global current_record
    data = request.json
    row = data.get('row', {})
    orig = data.get('original')
    row_list = [row.get(h, '') for h in headers]
    if orig is not None:
        if orig in keep_rows:
            keep_rows.remove(orig)
        if orig in skip_rows:
            skip_rows.remove(orig)
    if row_list in skip_rows:
        skip_rows.remove(row_list)
    keep_rows.append(row_list)
    current_record = None
    save_state()
    return jsonify(keep_rows=keep_rows, skip_rows=skip_rows)

@app.route('/download')
def download():
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    for r in keep_rows:
        writer.writerow(r)
    res = make_response(out.getvalue())
    fname = f"kept_contacts_{datetime.now().strftime('%Y%m%d')}.csv"
    res.headers['Content-Disposition'] = f'attachment; filename={fname}'
    res.headers['Content-Type'] = 'text/csv'
    return res

@app.route('/delete_state', methods=['POST'])
def delete_state():
    global all_rows, headers, keep_rows, skip_rows, current_record
    all_rows.clear()
    headers.clear()
    keep_rows.clear()
    skip_rows.clear()
    current_record = None
    if os.path.exists(STATE_FILE):
        os.remove(STATE_FILE)
    return jsonify(message="State file deleted.")

if __name__ == '__main__':
    app.run(port=5000, debug=True)